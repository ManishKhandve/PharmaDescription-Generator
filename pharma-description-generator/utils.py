"""
Utilities module for Excel file handling, batch processing, and data management.
Handles large files efficiently with partial saves and error recovery.
"""

import pandas as pd
import os
import logging
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Handles Excel file operations including reading input and writing formatted output.
    """
    
    @staticmethod
    def read_input_file(file_path: str) -> tuple[list[dict], pd.DataFrame]:
        """
        Read product names from the first column and preserve all original data.
        
        Args:
            file_path (str): Path to input Excel file
            
        Returns:
            tuple: (List of product names, Original DataFrame with all data)
            
        Raises:
            Exception: If file cannot be read or is invalid
        """
        try:
            # Input validation
            if not file_path or not isinstance(file_path, str):
                raise ValueError(f"Invalid file path: {file_path}")
            
            file_path = str(file_path).strip()
            if not file_path:
                raise ValueError("Empty file path provided")
            
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if not file_path.lower().endswith(('.xlsx', '.xls')):
                raise ValueError(f"Invalid file format. Expected .xlsx or .xls, got: {file_path}")
            
            # Read Excel file with all data
            try:
                df = pd.read_excel(file_path, header=0, engine='openpyxl')
            except Exception as e:
                # Fallback to xlrd for older files
                try:
                    df = pd.read_excel(file_path, header=0, engine='xlrd')
                except Exception as e2:
                    raise Exception(f"Could not read Excel file with any engine. openpyxl error: {str(e)}, xlrd error: {str(e2)}")
            
            if df.empty:
                raise ValueError("Excel file is empty")
            
            if len(df.columns) == 0:
                raise ValueError("Excel file has no columns")
            
            # Get first column values, remove NaN and empty strings
            try:
                first_column = df.iloc[:, 0]
                products = first_column.dropna().astype(str).tolist()
                products = [str(p).strip() for p in products if str(p).strip() and str(p).strip().lower() != 'nan']
            except Exception as e:
                logger.error(f"Error processing first column: {str(e)}")
                raise ValueError(f"Could not extract product names from first column: {str(e)}")

            # Check for 'Category' column (case-insensitive)
            category_col = None
            for col in df.columns:
                if str(col).strip().lower() == 'category':
                    category_col = col
                    break

            product_info = []
            for idx, pname in enumerate(products):
                cat = None
                if category_col is not None:
                    try:
                        cat_val = df.at[idx, category_col]
                        if pd.notna(cat_val):
                            cat = str(cat_val).strip()
                    except Exception:
                        cat = None
                product_info.append({"product_name": pname, "category": cat})

            if not product_info:
                raise ValueError("No valid product names found in first column")

            # Remove any existing description columns from the original DataFrame to avoid conflicts
            # This ensures we generate fresh descriptions regardless of pre-existing ones
            description_keywords = [
                'description', 'short description', 'long description', 'short_description', 'long_description',
                'desc', 'short desc', 'long desc', 'product description', 'item description',
                'details', 'product details', 'summary', 'overview'
            ]
            
            columns_to_remove = []
            for col in df.columns:
                col_name_lower = str(col).strip().lower()
                if any(keyword in col_name_lower for keyword in description_keywords):
                    columns_to_remove.append(col)
                    logger.info(f"Ignoring existing description column: '{col}'")
            
            if columns_to_remove:
                df = df.drop(columns=columns_to_remove)
                logger.info(f"Removed {len(columns_to_remove)} existing description columns")

            logger.info(f"Successfully read {len(product_info)} products from {file_path}")
            return product_info, df
            
        except Exception as e:
            logger.error(f"Error reading input file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def create_output_bytes(results: List[Dict[str, Any]], original_data: pd.DataFrame = None) -> bytes:
        """
        Create formatted Excel output file in memory and return as bytes.
        
        Args:
            results (List[Dict]): List of processed product results
            original_data (pd.DataFrame): Original input data to preserve
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            # Input validation
            if not results or not isinstance(results, list):
                raise ValueError(f"Invalid results data: {type(results)}")
            
            if len(results) == 0:
                raise ValueError("No results to write to file")
            
            # Validate results structure
            for i, result in enumerate(results):
                if not isinstance(result, dict):
                    logger.warning(f"Invalid result at index {i}: {type(result)}")
                    continue
                if 'product_name' not in result:
                    logger.warning(f"Missing product_name in result {i}")
            
            if original_data is not None and not original_data.empty:
                try:
                    # Create enhanced output with original data preserved
                    output_df = original_data.copy()
                    
                    # Create a mapping of product names to results
                    results_map = {}
                    for result in results:
                        try:
                            product_name = str(result.get('product_name', '')).strip()
                            if product_name:
                                results_map[product_name] = result
                        except Exception as e:
                            logger.warning(f"Error processing result: {str(e)}")
                    
                    # Create description lists with error handling
                    short_descriptions = []
                    long_descriptions = []
                    
                    for idx, row in output_df.iterrows():
                        try:
                            product_name = str(row.iloc[0]).strip() if len(row) > 0 else ''
                            result = results_map.get(product_name, {})
                            
                            short_desc = result.get('short_description', '')
                            long_desc = result.get('long_description', '')
                            
                            # Clean descriptions and format short description with HTML
                            if short_desc:
                                short_desc = str(short_desc).replace('*', '').replace('_', '').strip()
                                # Convert to HTML format with proper bullet points
                                lines = short_desc.split('\n')
                                html_lines = []
                                for line in lines:
                                    try:
                                        line = str(line).strip()
                                        if line.startswith('*'):
                                            line = line[1:].strip()
                                        elif line.startswith('- '):
                                            line = line[2:].strip()
                                        elif line.startswith('• '):
                                            line = line[2:].strip()
                                        
                                        if line:  # Only add non-empty lines
                                            html_lines.append(f"<li>{line}</li>")
                                    except Exception as e:
                                        logger.warning(f"Error processing bullet line: {str(e)}")
                                        if line:
                                            html_lines.append(f"<li>{str(line).replace('*', '').strip()}</li>")
                                
                                if html_lines:
                                    short_desc = f"<ul>{''.join(html_lines)}</ul>"
                                else:
                                    short_desc = short_desc  # Keep original if no bullet points found
                            
                            if long_desc:
                                long_desc = str(long_desc).replace('*', '').replace('_', '').strip()
                            
                            short_descriptions.append(short_desc)
                            long_descriptions.append(long_desc)
                            
                        except Exception as e:
                            logger.warning(f"Error processing row {idx}: {str(e)}")
                            short_descriptions.append('')
                            long_descriptions.append('')
                    
                    # Add new columns with descriptions - guaranteed no overwrite
                    output_df['Short Description'] = short_descriptions
                    output_df['Long Description'] = long_descriptions
                    
                except Exception as e:
                    logger.error(f"Error processing original data: {str(e)}")
                    # Fallback to simple format
                    output_df = ExcelHandler._create_simple_output(results)
                    
            else:
                # Fallback to simple format if no original data
                try:
                    data = []
                    for result in results:
                        try:
                            data.append({
                                'Product Name': str(result.get('product_name', '')),
                                'Short Description': str(result.get('short_description', '')).replace('*', ''),
                                'Long Description': str(result.get('long_description', '')).replace('*', '')
                            })
                        except Exception as e:
                            logger.warning(f"Error processing result: {str(e)}")
                            data.append({
                                'Product Name': 'Error',
                                'Short Description': '',
                                'Long Description': ''
                            })
                    output_df = pd.DataFrame(data)
                except Exception as e:
                    logger.error(f"Error creating simple DataFrame: {str(e)}")
                    output_df = pd.DataFrame([{
                        'Product Name': 'Error',
                        'Short Description': '',
                        'Long Description': ''
                    }])
            
            # Create workbook and worksheet in memory with error handling
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Product Descriptions"
                
                # Add DataFrame to worksheet
                for r in dataframe_to_rows(output_df, index=False, header=True):
                    ws.append(r)
                
                # Format the worksheet
                ExcelHandler._format_worksheet(ws)
                
                # Save to BytesIO object
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                logger.info(f"Successfully created Excel file in memory with {len(results)} products")
                return excel_buffer.getvalue()
                
            except Exception as e:
                logger.error(f"Error creating Excel workbook: {str(e)}")
                # Try minimal fallback
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = 'Product Name'
                    ws['B1'] = 'Short Description'
                    ws['C1'] = 'Long Description'
                    
                    for i, result in enumerate(results, 2):
                        ws[f'A{i}'] = str(result.get('product_name', ''))
                        ws[f'B{i}'] = str(result.get('short_description', '')).replace('*', '')
                        ws[f'C{i}'] = str(result.get('long_description', '')).replace('*', '')
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    return excel_buffer.getvalue()
                    
                except Exception as e2:
                    logger.critical(f"Fallback Excel creation failed: {str(e2)}")
                    raise Exception(f"Failed to create Excel file: {str(e)}")
            
        except Exception as e:
            logger.error(f"Critical error in create_output_bytes: {str(e)}")
            raise
    
    @staticmethod
    def create_output_file(results: List[Dict[str, Any]], output_path: str, original_data: pd.DataFrame = None) -> str:
        """
        Create formatted Excel output file with product descriptions and original data.
        
        Args:
            results (List[Dict]): List of processed product results
            output_path (str): Path for output file
            original_data (pd.DataFrame): Original input data to preserve
            
        Returns:
            str: Path to created output file
        """
        try:
            # Input validation
            if not results or not isinstance(results, list):
                raise ValueError(f"Invalid results data: {type(results)}")
            
            if not output_path or not isinstance(output_path, str):
                raise ValueError(f"Invalid output path: {output_path}")
            
            if len(results) == 0:
                raise ValueError("No results to write to file")
            
            # Ensure output directory exists
            try:
                output_dir = os.path.dirname(output_path)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
            except Exception as e:
                logger.warning(f"Error creating output directory: {str(e)}")
            
            if original_data is not None and not original_data.empty:
                try:
                    # Create enhanced output with original data preserved
                    output_df = original_data.copy()
                    
                    # Create a mapping of product names to results for faster lookup with fuzzy matching
                    results_map = {}
                    results_map_normalized = {}
                    for result in results:
                        try:
                            product_name = str(result.get('product_name', '')).strip()
                            if product_name:
                                results_map[product_name] = result
                                # Create normalized version for fuzzy matching
                                normalized_name = product_name.lower().replace(' ', '').replace('-', '').replace('_', '')
                                results_map_normalized[normalized_name] = result
                        except Exception as e:
                            logger.warning(f"Error processing result mapping: {str(e)}")
                    
                    # Find available column positions for descriptions - improved logic
                    total_cols = len(output_df.columns)
                    
                    # Always place descriptions in predictable locations to avoid missed columns
                    short_desc_col = total_cols      # Always append as new column
                    long_desc_col = total_cols + 1   # Always append as new column
                    
                    # Add description data with guaranteed column placement
                    short_descriptions = []
                    long_descriptions = []
                    
                    for idx, row in output_df.iterrows():
                        try:
                            product_name = str(row.iloc[0]).strip() if len(row) > 0 else ''
                            
                            # Try exact match first
                            result = results_map.get(product_name, {})
                            
                            # If no exact match, try fuzzy matching
                            if not result and product_name:
                                normalized_input = product_name.lower().replace(' ', '').replace('-', '').replace('_', '')
                                result = results_map_normalized.get(normalized_input, {})
                                
                                # If still no match, try partial matching
                                if not result:
                                    for norm_key, norm_result in results_map_normalized.items():
                                        if norm_key in normalized_input or normalized_input in norm_key:
                                            result = norm_result
                                            logger.info(f"Found partial match for '{product_name}' -> '{norm_result.get('product_name', '')}'")
                                            break
                            
                            # If no result found, log the missing entry
                            if not result and product_name:
                                logger.warning(f"No description found for product: '{product_name}'")
                            
                            short_desc = result.get('short_description', '')
                            long_desc = result.get('long_description', '')
                            
                            # Clean descriptions and format short description with HTML
                            if short_desc:
                                short_desc = str(short_desc).replace('*', '').replace('_', '').strip()
                                # Convert to HTML format with proper bullet points
                                lines = short_desc.split('\n')
                                html_lines = []
                                for line in lines:
                                    try:
                                        line = str(line).strip()
                                        if line.startswith('*'):
                                            line = line[1:].strip()
                                        elif line.startswith('- '):
                                            line = line[2:].strip()
                                        elif line.startswith('• '):
                                            line = line[2:].strip()
                                        
                                        if line:  # Only add non-empty lines
                                            html_lines.append(f"<li>{line}</li>")
                                    except Exception as e:
                                        logger.warning(f"Error processing bullet line: {str(e)}")
                                        if line:
                                            html_lines.append(f"<li>{str(line).replace('*', '').strip()}</li>")
                                
                                if html_lines:
                                    short_desc = f"<ul>{''.join(html_lines)}</ul>"
                                else:
                                    short_desc = short_desc  # Keep original if no bullet points found
                            
                            if long_desc:
                                long_desc = str(long_desc).replace('*', '').replace('_', '').strip()
                            
                            short_descriptions.append(short_desc)
                            long_descriptions.append(long_desc)
                            
                        except Exception as e:
                            logger.warning(f"Error processing row {idx}: {str(e)}")
                            short_descriptions.append('')
                            long_descriptions.append('')
                    
                    # Add new columns with descriptions - guaranteed no overwrite
                    output_df['Short Description'] = short_descriptions
                    output_df['Long Description'] = long_descriptions
                    
                except Exception as e:
                    logger.error(f"Error processing original data: {str(e)}")
                    # Fallback to simple format
                    output_df = ExcelHandler._create_simple_output(results)
                
            else:
                # Fallback to simple format if no original data
                try:
                    data = []
                    for result in results:
                        try:
                            data.append({
                                'Product Name': str(result.get('product_name', '')),
                                'Short Description': str(result.get('short_description', '')).replace('*', ''),
                                'Long Description': str(result.get('long_description', '')).replace('*', '')
                            })
                        except Exception as e:
                            logger.warning(f"Error processing result: {str(e)}")
                            data.append({
                                'Product Name': 'Error',
                                'Short Description': '',
                                'Long Description': ''
                            })
                    output_df = pd.DataFrame(data)
                except Exception as e:
                    logger.error(f"Error creating simple DataFrame: {str(e)}")
                    output_df = pd.DataFrame([{
                        'Product Name': 'Error',
                        'Short Description': '',
                        'Long Description': ''
                    }])
            
            # Create workbook and worksheet with error handling
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Product Descriptions"
                
                # Add DataFrame to worksheet
                for r in dataframe_to_rows(output_df, index=False, header=True):
                    ws.append(r)
                
                # Format the worksheet
                ExcelHandler._format_worksheet(ws)
                
                # Save workbook
                wb.save(output_path)
                
                logger.info(f"Successfully created output file: {output_path}")
                return output_path
                
            except Exception as e:
                logger.error(f"Error creating Excel workbook: {str(e)}")
                # Try minimal fallback
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws['A1'] = 'Product Name'
                    ws['B1'] = 'Short Description'
                    ws['C1'] = 'Long Description'
                    
                    for i, result in enumerate(results, 2):
                        ws[f'A{i}'] = str(result.get('product_name', ''))
                        ws[f'B{i}'] = str(result.get('short_description', '')).replace('*', '')
                        ws[f'C{i}'] = str(result.get('long_description', '')).replace('*', '')
                    
                    wb.save(output_path)
                    logger.info(f"Created fallback Excel file: {output_path}")
                    return output_path
                    
                except Exception as e2:
                    logger.critical(f"Fallback Excel creation failed: {str(e2)}")
                    raise Exception(f"Failed to create Excel file: {str(e)}")
            
        except Exception as e:
            logger.error(f"Error creating output file: {str(e)}")
            raise Exception(f"Failed to create output file: {str(e)}")
    
    @staticmethod
    def _create_simple_output(results: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Create a simple DataFrame output as fallback.
        
        Args:
            results (List[Dict]): List of processed product results
            
        Returns:
            pd.DataFrame: Simple formatted DataFrame
        """
        try:
            data = []
            for result in results:
                try:
                    data.append({
                        'Product Name': str(result.get('product_name', '')),
                        'Short Description': str(result.get('short_description', '')).replace('*', ''),
                        'Long Description': str(result.get('long_description', '')).replace('*', '')
                    })
                except Exception as e:
                    logger.warning(f"Error processing result: {str(e)}")
                    data.append({
                        'Product Name': 'Error',
                        'Short Description': '',
                        'Long Description': ''
                    })
            return pd.DataFrame(data)
        except Exception as e:
            logger.error(f"Error creating simple output: {str(e)}")
            return pd.DataFrame([{
                'Product Name': 'Error',
                'Short Description': '',
                'Long Description': ''
            }])

    @staticmethod
    def _format_worksheet(worksheet):
        """
        Apply professional formatting to the Excel worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header formatting
            for cell in worksheet[1]:
                try:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                except Exception as e:
                    logger.warning(f"Error formatting header cell: {str(e)}")
            
            # Set column widths
            column_widths = {
                'A': 30,  # Product Name
                'B': 50,  # Short Description
                'C': 70   # Long Description
            }
            
            for column, width in column_widths.items():
                try:
                    worksheet.column_dimensions[column].width = width
                except Exception as e:
                    logger.warning(f"Error setting column width for {column}: {str(e)}")
            
            # Set text wrapping for description columns
            try:
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row[1:]:  # Skip product name column
                        try:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        except Exception as e:
                            logger.warning(f"Error setting text wrapping: {str(e)}")
            except Exception as e:
                logger.warning(f"Error setting text wrapping for rows: {str(e)}")
                
        except Exception as e:
            logger.error(f"Error formatting worksheet: {str(e)}")
    
    @staticmethod
    def split_large_file(results: List[Dict[str, Any]], max_rows: int = 5000) -> List[List[Dict[str, Any]]]:
        """
        Split large result sets into smaller chunks for multiple output files.
        
        Args:
            results (List[Dict]): Complete results list
            max_rows (int): Maximum rows per file
            
        Returns:
            List[List[Dict]]: List of result chunks
        """
        try:
            # Input validation
            if not results or not isinstance(results, list):
                logger.warning(f"Invalid results for splitting: {type(results)}")
                return []
            
            if not isinstance(max_rows, int) or max_rows <= 0:
                logger.warning(f"Invalid max_rows: {max_rows}, using default 5000")
                max_rows = 5000
            
            chunks = []
            for i in range(0, len(results), max_rows):
                try:
                    chunk = results[i:i + max_rows]
                    if chunk:  # Only add non-empty chunks
                        chunks.append(chunk)
                except Exception as e:
                    logger.warning(f"Error creating chunk {i//max_rows}: {str(e)}")
            
            logger.info(f"Split {len(results)} results into {len(chunks)} chunks")
            return chunks
            
        except Exception as e:
            logger.error(f"Error splitting file: {str(e)}")
            return [results] if results else []


class ProgressTracker:
    """
    Handles progress tracking and partial saves during processing.
    """
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize progress tracker.
        
        Args:
            output_dir (str): Directory to store partial saves
        """
        try:
            self.output_dir = str(output_dir) if output_dir else "output"
            self.results = []
            self.processed_count = 0
            self.total_count = 0
            
            # Create output directory if it doesn't exist
            os.makedirs(self.output_dir, exist_ok=True)
            logger.info(f"Initialized ProgressTracker with output dir: {self.output_dir}")
            
        except Exception as e:
            logger.error(f"Error initializing ProgressTracker: {str(e)}")
            self.output_dir = "output"
            self.results = []
            self.processed_count = 0
            self.total_count = 0
    
    def initialize(self, total_products: int):
        """
        Initialize tracking for a new batch.
        
        Args:
            total_products (int): Total number of products to process
        """
        try:
            self.total_count = max(0, int(total_products)) if total_products else 0
            self.processed_count = 0
            self.results = []
            logger.info(f"Initialized progress tracking for {self.total_count} products")
            
        except Exception as e:
            logger.error(f"Error initializing progress: {str(e)}")
            self.total_count = 0
            self.processed_count = 0
            self.results = []
    
    def update_progress(self, new_results: List[Dict[str, Any]]):
        """
        Update progress with new results and perform partial save.
        
        Args:
            new_results (List[Dict]): New batch of processed results
        """
        self.results.extend(new_results)
        self.processed_count = len(self.results)
        
        # Perform partial save every 100 products
        if self.processed_count % 100 == 0:
            self._partial_save()
        
        logger.info(f"Progress: {self.processed_count}/{self.total_count} products processed")
    
    def _partial_save(self):
        """
        Save current results to a temporary file.
        """
        try:
            # Note: Partial saves disabled in favor of in-memory processing
            logger.info(f"Partial save skipped - using in-memory processing: {self.processed_count} products")
        except Exception as e:
            logger.error(f"Partial save failed: {str(e)}")
    
    def get_progress_percentage(self) -> float:
        """
        Get current progress as percentage.
        
        Returns:
            float: Progress percentage (0-100)
        """
        if self.total_count == 0:
            return 0.0
        return (self.processed_count / self.total_count) * 100
    
    def get_results(self) -> List[Dict[str, Any]]:
        """
        Get all processed results.
        
        Returns:
            List[Dict]: Complete results list
        """
        return self.results.copy()


class FileValidator:
    """
    Validates input files and handles file-related errors.
    """
    
    @staticmethod
    def validate_input_file(file_path: str) -> Dict[str, Any]:
        """
        Validate input Excel file and return file information.
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            Dict[str, Any]: Validation result with file info
        """
        result = {
            'valid': False,
            'error': None,
            'product_count': 0,
            'file_size': 0
        }
        
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                result['error'] = "File does not exist"
                return result
            
            # Check file size
            file_size = os.path.getsize(file_path)
            result['file_size'] = file_size
            
            # Check if file is too large (>50MB)
            if file_size > 50 * 1024 * 1024:
                result['error'] = "File too large (maximum 50MB)"
                return result
            
            # Try to read file
            products, _ = ExcelHandler.read_input_file(file_path)  # Only need product count for validation
            result['product_count'] = len(products)
            
            # Check if too many products
            if len(products) > 50000:
                result['error'] = f"Too many products ({len(products)}). Maximum 50,000 supported."
                return result
            
            result['valid'] = True
            return result
            
        except Exception as e:
            result['error'] = str(e)
            return result
    
    @staticmethod
    def get_safe_filename(original_name: str) -> str:
        """
        Generate safe filename for output files.
        
        Args:
            original_name (str): Original filename
            
        Returns:
            str: Safe filename
        """
        # Remove extension and clean name
        name = os.path.splitext(original_name)[0]
        safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        
        if not safe_name:
            safe_name = "output"
        
        return f"{safe_name}_descriptions.xlsx"


def estimate_processing_time(product_count: int, model_type: str) -> Dict[str, Any]:
    """
    Estimate processing time based on product count and model type.
    
    Args:
        product_count (int): Number of products to process
        model_type (str): LLM model type ('mistral' or 'gemini')
        
    Returns:
        Dict[str, Any]: Time estimates and recommendations
    """
    # Average processing times per product (in seconds)
    base_times = {
        'mistral': 3.0,  # Faster API
        'gemini': 4.0    # Slightly slower due to rate limits
    }
    
    base_time = base_times.get(model_type.lower(), 3.5)
    
    # Account for batch processing efficiency
    batch_factor = 0.7  # 30% efficiency gain from batching
    
    estimated_seconds = (product_count * base_time) * batch_factor
    
    # Convert to human-readable format (always show h, m, s)
    total_seconds = int(estimated_seconds)
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    time_str = f"{hours}h {minutes}m {seconds}s"
    
    return {
        'estimated_seconds': estimated_seconds,
        'formatted_time': time_str,
        'recommendation': _get_processing_recommendation(product_count)
    }


def _get_processing_recommendation(product_count: int) -> str:
    """
    Get processing recommendation based on product count.
    
    Args:
        product_count (int): Number of products
        
    Returns:
        str: Recommendation message
    """
    if product_count < 100:
        return "Small batch - processing should complete quickly."
    elif product_count < 1000:
        return "Medium batch - estimated completion in a few minutes."
    elif product_count < 5000:
        return "Large batch - processing may take 15-30 minutes. Results will be saved periodically."
    else:
        return "Very large batch - processing may take 1+ hours. Consider splitting into smaller files."
